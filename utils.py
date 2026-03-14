"""
Exporter module - saves scraped data to CSV, Excel, or JSON
"""
import json
import logging
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

EXPORT_COLUMNS = ["title", "price", "phone", "seller", "location", "url", "description", "date"]


def get_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def export_csv(data: list[dict], output_path: str = None) -> str:
    import csv
    if not output_path:
        output_path = f"milanuncios_export_{get_timestamp()}.csv"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)
    logger.info(f"Exported {len(data)} records to {output_path}")
    return output_path


def export_excel(data: list[dict], output_path: str = None) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not output_path:
        output_path = f"milanuncios_export_{get_timestamp()}.xlsx"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Milanuncios Data"

    # Header styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1a1a2e")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["Title", "Price", "Phone", "Seller", "Location", "URL", "Description", "Date"]
    col_widths = [40, 14, 14, 20, 20, 50, 60, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Data rows with alternating colors
    light_fill = PatternFill("solid", start_color="F8F9FA")
    white_fill = PatternFill("solid", start_color="FFFFFF")
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, record in enumerate(data, 2):
        fill = light_fill if row_idx % 2 == 0 else white_fill
        row_data = [
            record.get("title", ""),
            record.get("price", ""),
            record.get("phone", ""),
            record.get("seller", ""),
            record.get("location", ""),
            record.get("url", ""),
            record.get("description", "")[:500] if record.get("description") else "",
            record.get("date", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = data_align
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)

        ws.row_dimensions[row_idx].height = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Export Summary"
    ws_summary["A1"].font = Font(bold=True, size=14, name="Arial")
    ws_summary["A3"] = "Total Listings:"
    ws_summary["B3"] = len(data)
    ws_summary["A4"] = "Export Date:"
    ws_summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A5"] = "Listings with Phone:"
    ws_summary["B5"] = f"=COUNTIF('Milanuncios Data'!C:C,\"?*\")"

    wb.save(output_path)
    logger.info(f"Exported {len(data)} records to {output_path}")
    return output_path


def export_json(data: list[dict], output_path: str = None) -> str:
    if not output_path:
        output_path = f"milanuncios_export_{get_timestamp()}.json"
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.info(f"Exported {len(data)} records to {output_path}")
    return output_path
